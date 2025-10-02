import streamlit as st
import sqlite3
import pandas as pd
from datetime import date
from io import BytesIO

# ReportLab untuk PDF
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

# Openpyxl untuk Excel
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# -------------------------
# Setup Database
# -------------------------
conn = sqlite3.connect("checklists.db", check_same_thread=False)
c = conn.cursor()

c.execute("""
CREATE TABLE IF NOT EXISTS items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    judul TEXT,
    area TEXT,
    assignedTo TEXT,
    tanggal TEXT,
    keterangan TEXT,
    ok INTEGER,
    rekomendasi TEXT
)
""")
conn.commit()

# -------------------------
# Helper Functions
# -------------------------
def get_items():
    return pd.read_sql("SELECT * FROM items", conn)

def add_item(judul, area, assignedTo, tanggal, keterangan, ok="T", rekomendasi=""):
    ok_val = 1 if ok == "Y" else 0
    c.execute(
        "INSERT INTO items (judul,area,assignedTo,tanggal,keterangan,ok,rekomendasi) VALUES (?,?,?,?,?,?,?)",
        (judul, area, assignedTo, tanggal, keterangan, ok_val, rekomendasi)
    )
    conn.commit()

def update_item(id, judul=None, area=None, assignedTo=None, tanggal=None, keterangan=None, ok=None, rekomendasi=None):
    if judul is not None:
        c.execute("UPDATE items SET judul=? WHERE id=?", (judul, id))
    if area is not None:
        c.execute("UPDATE items SET area=? WHERE id=?", (area, id))
    if assignedTo is not None:
        c.execute("UPDATE items SET assignedTo=? WHERE id=?", (assignedTo, id))
    if tanggal is not None:
        c.execute("UPDATE items SET tanggal=? WHERE id=?", (tanggal, id))
    if keterangan is not None:
        c.execute("UPDATE items SET keterangan=? WHERE id=?", (keterangan, id))
    if ok is not None:
        ok_val = 1 if ok == "Y" else 0
        c.execute("UPDATE items SET ok=? WHERE id=?", (ok_val, id))
    if rekomendasi is not None:
        c.execute("UPDATE items SET rekomendasi=? WHERE id=?", (rekomendasi, id))
    conn.commit()

def delete_item(id):
    c.execute("DELETE FROM items WHERE id=?", (id,))
    conn.commit()

def search_items_df(keyword):
    df = pd.read_sql(
        """
        SELECT * FROM items
        WHERE judul LIKE ? 
           OR area LIKE ? 
           OR assignedTo LIKE ? 
           OR tanggal LIKE ? 
           OR keterangan LIKE ? 
           OR rekomendasi LIKE ?
        """, 
        conn, params=tuple([f"%{keyword}%"] * 6)
    )
    return df

# -------------------------
# Streamlit Config & UI
# -------------------------
st.set_page_config(page_title="Audit & Checklist Tool", layout="wide")
st.title("📋 Audit & Checklist Tool")

# -------------------------
# Sidebar Form (Tambah Item)
# -------------------------
st.sidebar.header("➕ Tambah Item Baru")

with st.sidebar.form("form_tambah_item", clear_on_submit=True):
    judul = st.text_input("Hal yang Diperiksa")
    area = st.text_input("Area")
    assignedTo = st.text_input("Assigned To")
    tanggal = st.date_input("Tanggal", value=date.today())
    keterangan = st.text_area("Hasil Temuan")
    ok = st.radio("Check", ["Y", "T"], horizontal=True)
    rekomendasi = st.text_area("Rekomendasi")

    submitted = st.form_submit_button("💾 Simpan Item")
    if submitted:
        add_item(judul, area, assignedTo, str(tanggal), keterangan, ok, rekomendasi)
        st.success("✅ Item berhasil ditambahkan")
        st.rerun()

# -------------------------
# Tampilkan & Edit Items
# -------------------------
st.subheader("📌 Daftar Items")
items = get_items()

if not items.empty:
    # Filter dropdowns
    st.write("🔍 **Filter Data**")
    col_f1, col_f2, col_f3 = st.columns(3)
    with col_f1:
        filter_judul = st.selectbox("Filter Hal yang Diperiksa", ["(Semua)"] + sorted(items["judul"].dropna().unique().tolist()))
    with col_f2:
        filter_area = st.selectbox("Filter Area", ["(Semua)"] + sorted(items["area"].dropna().unique().tolist()))
    with col_f3:
        filter_assigned = st.selectbox("Filter Assigned To", ["(Semua)"] + sorted(items["assignedTo"].dropna().unique().tolist()))

    filtered_items = items.copy()
    if filter_judul != "(Semua)":
        filtered_items = filtered_items[filtered_items["judul"] == filter_judul]
    if filter_area != "(Semua)":
        filtered_items = filtered_items[filtered_items["area"] == filter_area]
    if filter_assigned != "(Semua)":
        filtered_items = filtered_items[filtered_items["assignedTo"] == filter_assigned]

    if filtered_items.empty:
        st.warning("⚠️ Tidak ada data sesuai filter")
    else:
        for _, row in filtered_items.iterrows():
            rid = int(row["id"])
            with st.expander(f"📝 {row['judul']} - {row['area']} - {row['assignedTo']} (ID: {rid})"):
                new_judul = st.text_input("Hal yang Diperiksa", row["judul"], key=f"judul-{rid}")
                new_area = st.text_input("Area", row["area"], key=f"area-{rid}")
                new_assigned = st.text_input("Assigned To", row["assignedTo"], key=f"assigned-{rid}")
                new_tanggal = st.date_input("Tanggal", value=pd.to_datetime(row["tanggal"]), key=f"tanggal-{rid}")
                new_keterangan = st.text_area("Hasil Temuan", row["keterangan"], key=f"ket-{rid}")
                new_ok = st.radio("Check", ["Y", "T"], index=0 if row["ok"] == 1 else 1, key=f"ok-{rid}")
                new_rekomendasi = st.text_area("Rekomendasi", row["rekomendasi"], key=f"rek-{rid}")

                col_u, col_d = st.columns(2)
                with col_u:
                    if st.button("💾 Update", key=f"update-{rid}"):
                        update_item(rid, judul=new_judul, area=new_area, assignedTo=new_assigned,
                                    tanggal=str(new_tanggal), keterangan=new_keterangan, ok=new_ok,
                                    rekomendasi=new_rekomendasi)
                        st.success("✅ Item updated!")
                        st.rerun()
                with col_d:
                    if st.button("🗑️ Hapus", key=f"delete-{rid}"):
                        delete_item(rid)
                        st.warning("❌ Item dihapus")
                        st.rerun()
else:
    st.info("Belum ada data. Tambahkan item baru dari sidebar.")

# -------------------------
# Export Excel & PDF
# -------------------------
items = get_items()
if not items.empty:
    st.subheader("📤 Export Data")

    export_data = items.rename(columns={
        "judul": "Hal yang Diperiksa",
        "area": "Area",
        "assignedTo": "Assigned To",
        "tanggal": "Tanggal",
        "keterangan": "Hasil Temuan",
        "ok": "Check",
        "rekomendasi": "Rekomendasi"
    }).fillna({"Rekomendasi": "", "Hasil Temuan": ""})

    if "id" in export_data.columns:
        export_data = export_data.drop(columns=["id"])

    export_data.insert(0, "No", range(1, len(export_data) + 1))
    export_data["Check"] = export_data["Check"].apply(lambda x: "✓" if int(x) == 1 else "✗")

    # --- Excel ---
    excel_buffer = BytesIO()
    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
        export_data.to_excel(writer, sheet_name="AuditChecklist", index=False)
        worksheet = writer.sheets["AuditChecklist"]

        # Atur lebar kolom
        col_widths = {
            "No": 5,
            "Hal yang Diperiksa": 30,
            "Area": 25,
            "Assigned To": 20,
            "Tanggal": 15,
            "Hasil Temuan": 50,
            "Check": 10,
            "Rekomendasi": 50,
        }
        for col_idx, col_name in enumerate(export_data.columns, 1):
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = col_widths.get(col_name, 20)

        # Border tipis
        thin_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000")
        )

        # Alignment
        wrap_alignment = Alignment(wrap_text=True, vertical="top")
        center_alignment = Alignment(horizontal="center", vertical="center")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for row_idx, row in enumerate(
            worksheet.iter_rows(min_row=2, max_row=worksheet.max_row,
                                min_col=1, max_col=worksheet.max_column),
            start=2
        ):
            check_value = worksheet.cell(row=row_idx, column=export_data.columns.get_loc("Check") + 1).value
            row_fill = green_fill if check_value == "✓" else red_fill

            max_lines = 1
            for cell in row:
                col_header = worksheet.cell(row=1, column=cell.column).value
                if col_header in ["Hasil Temuan", "Rekomendasi"]:
                    cell.alignment = wrap_alignment
                elif col_header == "Check":
                    cell.alignment = center_alignment   # ✅ Tengah untuk Check
                else:
                    cell.alignment = Alignment(vertical="top")

                cell.fill = row_fill
                cell.border = thin_border

                if col_header in ["Hasil Temuan", "Rekomendasi"] and cell.value:
                    max_lines = max(max_lines, str(cell.value).count("\n") + 1)
            worksheet.row_dimensions[row_idx].height = 15 * max_lines

    excel_buffer.seek(0)
    st.download_button(
        label="⬇️ Download Excel",
        data=excel_buffer,
        file_name="audit_checklist.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # --- PDF ---
    pdf_buffer = BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()
    normal_style = styles["Normal"]
    normal_style.fontSize = 8
    normal_style.leading = 10
    normal_style.wordWrap = "CJK"

    elements.append(Paragraph("📋 Audit & Checklist Report", styles["Title"]))
    elements.append(Spacer(1, 12))

    table_data = [export_data.columns.tolist()]
    row_styles = []

    for row_idx, r in enumerate(export_data.itertuples(index=False), start=1):
        row_data = []
        for col, v in zip(export_data.columns, r):
            v = str(v) if pd.notna(v) else ""
            if col in ["Hasil Temuan", "Rekomendasi"]:
                row_data.append(Paragraph(v.replace("\n", "<br/>"), normal_style))
            else:
                row_data.append(Paragraph(v, normal_style))
        table_data.append(row_data)

        if r.Check == "✓":
            row_styles.append(("BACKGROUND", (0, row_idx), (-1, row_idx), colors.HexColor("#C6EFCE")))
        else:
            row_styles.append(("BACKGROUND", (0, row_idx), (-1, row_idx), colors.HexColor("#FFC7CE")))

    table = Table(
        table_data,
        repeatRows=1,
        colWidths=[25, 120, 100, 90, 70, 200, 40, 150]
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4CAF50")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 1), (-1, -1), "TOP"),
        ("ALIGN", (0, 1), (0, -1), "CENTER"),   # No
        ("ALIGN", (6, 1), (6, -1), "CENTER"),   # ✅ Tengah untuk Check
    ] + row_styles))
    elements.append(table)
    doc.build(elements)
    pdf_buffer.seek(0)

    st.download_button(
        label="⬇️ Download PDF",
        data=pdf_buffer,
        file_name="audit_checklist.pdf",
        mime="application/pdf"
    )

# -------------------------
# Search Box
# -------------------------
st.subheader("🔎 Cari Data")
search_keyword = st.text_input("Masukkan kata kunci (misal: 'helm', 'lapangan')")

if search_keyword:
    results = search_items_df(search_keyword)
    if not results.empty:
        st.success(f"Ditemukan {len(results)} hasil untuk '{search_keyword}'")
        results_display = results.drop(columns=["id"]) if "id" in results.columns else results
        st.dataframe(results_display, use_container_width=True)
    else:
        st.warning("Tidak ada data ditemukan.")